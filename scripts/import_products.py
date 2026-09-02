#!/usr/bin/env python3
"""
Generate lib/pharma.ts from "Jaydev Group - Product Sheet (Final)".

Replaces scripts/import_pharma.py, which read an older workbook that encoded
section as an inline header row and therapeutic segment as a CELL FILL COLOUR.
The new sheet is plainly tabular - one tab per category, one column per field -
so none of that inference is needed any more.

Tab -> section:
    Pharma Intermediates      -> intermediates          (+ the API it feeds)
    APIs - Human              -> apis-human             (+ therapeutic segment)
    APIs - Veterinary         -> apis-veterinary
    Ingredients & Excipients  -> ingredients-excipients (+ ingredient type)

Carried over from the old importer, because the sheet does not hold them:
  - data/pharma-short-names.json   trade names shown on cards, keyed by CAS
  - data/pharma-cas-overrides.json multi-form CAS (e.g. phytoene cis/trans)

Still deliberately absent rather than guessed: pharmacopoeia grade, DMF, CEP,
GMP status, packaging, manufacturers. A "-" in the sheet means unverified.

Usage:  python3 scripts/import_products.py <workbook.xlsx>
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "lib" / "pharma.ts"
SHORT_NAMES = ROOT / "data" / "pharma-short-names.json"
CAS_OVERRIDES = ROOT / "data" / "pharma-cas-overrides.json"

# Order matters: it is the order the facets render in, and the order used to
# resolve a product that appears on two tabs (earlier tab loses).
TABS = [
    ("APIs - Human", "apis-human", "APIs - Human"),
    ("APIs - Veterinary", "apis-veterinary", "APIs - Veterinary"),
    ("Pharma Intermediates", "intermediates", "Pharma Intermediates"),
    ("Ingredients & Excipients", "ingredients-excipients", "Ingredients & Excipients"),
]

CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")

# Optional columns. None of these exist in the sheet today - the pharma pages
# say "confirmed per enquiry" precisely because the source cannot answer them.
# Add any of these headers to a pharma tab and the value flows straight through
# to the product page and its Product schema; leave them out and nothing
# changes. Multi-value cells split on ";".
OPTIONAL_TEXT = {
    "Description": "description",
    "Grade": "grade",
    "Pharmacopoeia": "pharmacopoeia",
    "DMF": "dmf",
    "CEP": "cep",
    "GMP": "gmp",
    "MOQ": "moq",
}
OPTIONAL_LIST = {
    "Packaging": "packaging",
    "Applications": "applications",
}


def cell(row, header):
    v = (row.get(header) or "").strip()
    return "" if v in {"-", "n/a", "N/A", "TBC"} else v


def slug(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s[:70] or "item"


def cas_valid(cas: str) -> bool:
    """Official CAS check digit: sum of digit x position from the right, mod 10."""
    if not CAS_RE.match(cas):
        return False
    body, check = cas.replace("-", "")[:-1], int(cas[-1])
    return sum(int(d) * (i + 1) for i, d in enumerate(reversed(body))) % 10 == check


def rows(ws):
    data = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
    data = [r for r in data if any(r)]
    hdr, body = data[0], data[1:]
    return [{k: v for k, v in zip(hdr, r) if k} for r in body]


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python3 scripts/import_products.py <workbook.xlsx>")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"Workbook not found: {src}")
    wb = openpyxl.load_workbook(src, data_only=True)

    shorts, cas_forms = {}, {}
    if SHORT_NAMES.exists():
        shorts = {
            k: v.strip()
            for k, v in json.loads(SHORT_NAMES.read_text()).items()
            if not k.startswith("_") and isinstance(v, str) and v.strip()
        }
    if CAS_OVERRIDES.exists():
        for k, v in json.loads(CAS_OVERRIDES.read_text()).items():
            if not k.startswith("_") and isinstance(v, list) and v:
                cas_forms[k] = v

    items, seen, bad_cas, dupes = [], {}, [], []
    for tab, section, _ in TABS:
        if tab not in wb.sheetnames:
            sys.exit(f"Missing tab: {tab}")
        for r in rows(wb[tab]):
            name = r["Product"].strip()
            if not name:
                continue
            pid = slug(name)

            # The sheet's summary says Prednisolone moved to APIs - Human, but the
            # Intermediates row was never deleted, so it lands on two tabs. Earlier
            # tabs in TABS win; the loser is reported, never silently dropped.
            if pid in seen:
                dupes.append((name, seen[pid], tab))
                continue

            cas = (r.get("CAS") or "").strip()
            cas = "" if cas == "-" else cas
            if cas and not cas_valid(cas):
                bad_cas.append((name, cas))
                continue

            seg = (r.get("Category / Segment") or "").strip()
            seg = "" if seg == "-" else seg
            notes = (r.get("Notes") or "").strip()

            item = {"id": pid, "name": name, "section": section}
            if cas:
                item["cas"] = cas
            if pid in cas_forms:
                item["casForms"] = cas_forms[pid]
            if cas and cas in shorts:
                item["shortName"] = shorts[cas]
            if section == "apis-human" and seg:
                item["therapeuticSegment"] = seg
            if section == "ingredients-excipients" and seg:
                item["ingredientType"] = seg
            if section == "intermediates":
                api = (r.get("API") or "").strip()
                if api and api != "-":
                    item["forApi"] = api
            if "investigational" in notes.lower():
                item["investigational"] = True

            for header, key in OPTIONAL_TEXT.items():
                v = cell(r, header)
                if v:
                    item[key] = v
            for header, key in OPTIONAL_LIST.items():
                v = cell(r, header)
                if v:
                    parts = [x.strip() for x in v.split(";") if x.strip()]
                    if parts:
                        item[key] = parts

            seen[pid] = tab
            items.append(item)

    # Emit in a stable order so the generated file diffs cleanly.
    order = {s: i for i, (_, s, _) in enumerate(TABS)}
    items.sort(key=lambda i: (order[i["section"]], i["name"].lower()))

    segments = sorted({i["therapeuticSegment"] for i in items if i.get("therapeuticSegment")})
    ing_types = sorted({i["ingredientType"] for i in items if i.get("ingredientType")})

    body = ",\n".join("  " + json.dumps(i, ensure_ascii=False) for i in items)
    sec_body = ",\n".join(
        f"  {{ id: {json.dumps(s)}, label: {json.dumps(l)} }}" for _, s, l in TABS
    )
    seg_body = ",\n".join(
        f"  {{ id: {json.dumps(slug(s))}, label: {json.dumps(s)} }}" for s in segments
    )
    ing_body = ",\n".join(
        f"  {{ id: {json.dumps(slug(s))}, label: {json.dumps(s)} }}" for s in ing_types
    )

    ts = f"""// ─────────────────────────────────────────────
//  GENERATED FILE - do not edit by hand.
//  Regenerate:  python3 scripts/import_products.py <workbook.xlsx>
//  Source:      "Jaydev Group - Product Sheet (Final)"
//
//  Pharmacopoeia grade, DMF, CEP and GMP status are NOT in the source
//  sheet and are deliberately absent here rather than guessed. The UI
//  renders "Grade & documentation on request" until they are supplied.
// ─────────────────────────────────────────────

export type PharmaSection =
{chr(10).join(f"  | {json.dumps(s)}" for _, s, _ in TABS)};

export type PharmaProduct = {{
  id: string;
  /** Full chemical name. Always searchable, even when shortName is shown. */
  name: string;
  /** Trade name shown on cards, from data/pharma-short-names.json. */
  shortName?: string;
  section: PharmaSection;
  cas?: string;
  /** Set when the product is sold in more than one isomeric form. */
  casForms?: {{ label: string; cas: string }}[];
  /** Human APIs only - the therapeutic filter axis. */
  therapeuticSegment?: string;
  /** Ingredients & Excipients only - nutraceutical / excipient / amino acid. */
  ingredientType?: string;
  /** Intermediates only - the finished API this feeds into. Buyers search by
   *  the target drug, so it is part of the search index. */
  forApi?: string;
  investigational?: boolean;

  /* ── Optional, absent until the source sheet carries them. Every one of
     these renders on the product page the moment it is supplied; until then
     the page states that it is confirmed per enquiry rather than guessing. ── */
  description?: string;
  /** Pharmacopoeial grade, e.g. "IP / BP / USP". */
  grade?: string;
  pharmacopoeia?: string;
  dmf?: string;
  cep?: string;
  gmp?: string;
  moq?: string;
  packaging?: string[];
  applications?: string[];
}};

/** What a card and the search index display. */
export const displayName = (p: PharmaProduct) => p.shortName ?? p.name;

export const PHARMA_SECTIONS: {{ id: PharmaSection; label: string }}[] = [
{sec_body},
];

export const THERAPEUTIC_SEGMENTS: {{ id: string; label: string }}[] = [
{seg_body},
];

export const INGREDIENT_TYPES: {{ id: string; label: string }}[] = [
{ing_body},
];

export const pharmaProducts: PharmaProduct[] = [
{body},
];

export const pharmaById = (id: string) => pharmaProducts.find((p) => p.id === id);
"""
    OUT.write_text(ts)

    counts = {}
    for i in items:
        counts[i["section"]] = counts.get(i["section"], 0) + 1
    print(f"wrote {OUT.relative_to(ROOT)}  ({len(items)} items)")
    for _, s, l in TABS:
        print(f"  {l:28} {counts.get(s, 0):4}")
    print(f"  {'with CAS':28} {sum(1 for i in items if i.get('cas')):4}")
    print(f"  {'segmented (human APIs)':28} {sum(1 for i in items if i.get('therapeuticSegment')):4}")
    print(f"  {'intermediates w/ target API':28} {sum(1 for i in items if i.get('forApi')):4}")
    print(f"  {'short names applied':28} {sum(1 for i in items if i.get('shortName')):4}")
    print(f"  {'investigational':28} {sum(1 for i in items if i.get('investigational')):4}")

    optional_keys = list(OPTIONAL_TEXT.values()) + list(OPTIONAL_LIST.values())
    have = {k: sum(1 for i in items if i.get(k)) for k in optional_keys}
    if any(have.values()):
        print("\n  optional fields supplied:")
        for k, n in have.items():
            if n:
                print(f"      {k:16} {n:4}/{len(items)}")
    else:
        print(f"\n  optional fields (grade, DMF, CEP, GMP, MOQ, packaging,")
        print(f"  applications, description): none supplied yet - pages show")
        print(f"  'confirmed per enquiry'. Add the columns to any pharma tab.")

    missing = [i["name"] for i in items if not i.get("cas")]
    if missing:
        print(f"\n  CAS still unverified ({len(missing)}) - left absent, never guessed:")
        for n in missing:
            print(f"      - {n}")
    unseg = [i["name"] for i in items if i["section"] == "apis-human" and not i.get("therapeuticSegment")]
    if unseg:
        print(f"\n  Human APIs with no therapeutic segment ({len(unseg)}) - assign in the sheet:")
        for n in unseg:
            print(f"      - {n}")
    if dupes:
        print(f"\n  Same product on two tabs ({len(dupes)}) - kept the first, dropped the rest:")
        for name, kept, dropped in dupes:
            print(f"      - {name}: kept [{kept}], dropped [{dropped}]")
    if bad_cas:
        print("\n  !! FAILED CAS CHECK DIGIT - not emitted:")
        for n, c in bad_cas:
            print(f"      - {n}: {c}")
        sys.exit(1)


if __name__ == "__main__":
    main()

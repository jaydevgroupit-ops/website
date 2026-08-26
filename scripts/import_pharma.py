#!/usr/bin/env python3
"""
Generate lib/pharma.ts from the Jaydev Group product spreadsheet.

The workbook encodes information in three places:
  1. Sheet1 rows, with SECTION HEADERS sitting inline in the API column
     ("Intermediates", "Ingredients", "Pharmaceutical Excipients", "Others").
  2. Therapeutic segment as the CELL FILL COLOUR of the API cell, matched
     against the colour legend in Sheet1 columns B/C.
  3. Per-segment tabs carrying indications and manufacturer tables.

Nothing is inferred beyond that. Fields the sheet does not contain
(pharmacopoeia grade, DMF, CEP, GMP, packaging) are left absent rather
than guessed.

Deliberately NOT emitted:
  - manufacturers  : internal sourcing research. Emitting it would publish
                     supplier names and plant locations in the JS bundle.
  - applications   : the sheet's Usage column is prose, so splitting it
                     produced misleading fragments. Dropped at source.

Usage:  python3 scripts/import_pharma.py [path/to/workbook.xlsx]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SRC = Path.home() / "Downloads" / "Jaydev Group Product list.xlsx"
OUT = ROOT / "lib" / "pharma.ts"
SHORT_NAMES = ROOT / "data" / "pharma-short-names.json"
CAS_OVERRIDES = ROOT / "data" / "pharma-cas-overrides.json"

SECTION_HEADERS = {
    "intermediates": "intermediates",
    "ingredients": "ingredients",
    "pharmaceutical excipients": "excipients",
    "others": "others",
}
SECTION_LABELS = [
    ("apis", "APIs"),
    ("intermediates", "Intermediates"),
    ("ingredients", "Nutraceutical Ingredients"),
    ("excipients", "Pharmaceutical Excipients"),
    ("others", "Fine Chemicals & Others"),
]
CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")


def slug(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s[:70] or "item"


def cas_valid(cas: str) -> bool:
    """Official CAS check digit: sum of digits x position from the right, mod 10."""
    if not CAS_RE.match(cas):
        return False
    body, check = cas.replace("-", "")[:-1], int(cas[-1])
    return sum(int(d) * (i + 1) for i, d in enumerate(reversed(body))) % 10 == check


def rgb(cell):
    """Fill colour as an (r,g,b) tuple, or None when unfilled."""
    try:
        v = cell.fill.start_color.rgb
    except AttributeError:
        return None
    if not isinstance(v, str) or len(v) != 8 or v == "00000000":
        return None
    return tuple(int(v[i:i + 2], 16) for i in (2, 4, 6))


def near(a, b, tol=12):
    return a and b and all(abs(x - y) <= tol for x, y in zip(a, b))


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.exists():
        sys.exit(f"Workbook not found: {src}")
    wb = openpyxl.load_workbook(src)          # styled read, for fill colours
    wbv = openpyxl.load_workbook(src, data_only=True)

    # Hand-written short trade names, keyed by CAS. Never overwritten here.
    cas_fix, cas_forms, dropped = {}, {}, set()
    if CAS_OVERRIDES.exists():
        raw = json.loads(CAS_OVERRIDES.read_text())
        dropped = set(raw.get("_removed", {}))
        for k, v in raw.items():
            if k.startswith("_"):
                continue
            if isinstance(v, str) and v.strip():
                cas_fix[k] = v.strip()
            elif isinstance(v, list) and v:
                # A product sold in more than one isomeric form.
                cas_forms[k] = v
                cas_fix[k] = v[0]["cas"]

    shorts = {}
    if SHORT_NAMES.exists():
        shorts = {
            k: v.strip()
            for k, v in json.loads(SHORT_NAMES.read_text()).items()
            if not k.startswith("_") and isinstance(v, str) and v.strip()
        }

    # ---- 1. colour legend -> segment ----
    ws = wb["Sheet1"]
    legend = []
    for r in range(2, 40):
        seg = ws.cell(r, 2).value
        if not seg or str(seg).strip().lower() == "segment":
            continue
        c = rgb(ws.cell(r, 3))
        if c:
            legend.append((c, str(seg).strip()))

    def norm(s):
        return re.sub(r"[^a-z0-9]", "", s.lower())

    legend_by_norm = {norm(name): name for _, name in legend}

    # ---- 2. per-segment tabs: indications, manufacturers, segment ----
    # The tab an API appears on is authoritative for its segment; some rows in
    # Sheet1 were never colour-filled (Anticoagulants, Respiratory/Steroids),
    # so colour alone would silently drop whole segments.
    uses, tab_segment = {}, {}
    for name in wb.sheetnames:
        if name == "Sheet1":
            continue
        seg_name = legend_by_norm.get(norm(name))
        if seg_name is None:                       # tab title differs from legend
            for lnorm, lname in legend_by_norm.items():
                if norm(name) and (norm(name) in lnorm or lnorm in norm(name)):
                    seg_name = lname
                    break
        rows = [
            ["" if c is None else str(c).strip() for c in r]
            for r in wbv[name].iter_rows(values_only=True)
        ]
        mode = None
        for row in rows:
            cells = [c for c in row if c]
            if not cells:
                continue
            head = " ".join(cells[:4]).lower()
            if "api" in head and "usage" in head:
                mode = "api"
                continue
            if "company" in head and "location" in head:
                mode = "mfr"
                continue
            if mode == "api" and len(cells) >= 3:
                uses.setdefault(cells[1].lower(), cells[2])
                if seg_name:
                    tab_segment.setdefault(cells[1].lower(), seg_name)
            # Manufacturer tables are read past deliberately - see header note.

    # ---- 3. Sheet1 rows -> products ----
    wsv = wbv["Sheet1"]
    section = "apis"
    items, seen, bad_cas = [], set(), []
    for r in range(2, ws.max_row + 1):
        raw = wsv.cell(r, 4).value
        if not raw:
            continue
        name = str(raw).strip()
        name = re.sub(r"^\(P\d+\)\s*", "", name)   # spreadsheet artifact, e.g. "(P1)11β,17α-..."
        key = name.lower()
        if key in SECTION_HEADERS:
            section = SECTION_HEADERS[key]
            continue
        if key == "api":
            continue

        cas = wsv.cell(r, 5).value
        cas = str(cas).strip() if cas else ""
        if not CAS_RE.match(cas):
            cas = ""                                  # never guess a CAS

        pid = slug(name)
        if pid in dropped:
            continue                                  # removed from the catalogue on purpose
        if not cas and pid in cas_fix:
            cas = cas_fix[pid]                        # hand-verified, see data/
        if cas and not cas_valid(cas):
            bad_cas.append((name, cas))
            cas = ""                                  # refuse to emit a bad CAS

        # Tab membership first (authoritative), colour as the fallback.
        segment = tab_segment.get(key)
        if not segment:
            fill = rgb(ws.cell(r, 4))
            if fill:
                for lc, lname in legend:
                    if near(fill, lc):
                        segment = lname
                        break

        # A few genuine APIs (Heparin, Enoxaparin, Budesonide, NAC) sit under
        # "Others" in Sheet1 but appear on a therapeutic tab. The tab wins.
        eff_section = "apis" if key in tab_segment else section

        item = {"id": pid, "name": name, "section": eff_section}
        if cas:
            item["cas"] = cas
            if pid in cas_forms:
                for f in cas_forms[pid]:
                    if not cas_valid(f["cas"]):
                        bad_cas.append((f"{name} ({f['label']})", f["cas"]))
                item["casForms"] = cas_forms[pid]
            short = shorts.get(cas)
            if short:
                item["shortName"] = short
        if eff_section == "apis" and segment:
            item["therapeuticSegment"] = segment

        # Usage text is read only to detect investigational status - never stored.
        usage = uses.get(key) or (str(wsv.cell(r, 6).value).strip() if wsv.cell(r, 6).value else "")
        if usage and re.search(r"\binvestigational\b", usage, re.I):
            item["investigational"] = True

        # The sheet repeats a couple of rows verbatim; keep the first.
        if item["id"] in seen:
            continue
        seen.add(item["id"])
        items.append(item)

    segments = sorted({i["therapeuticSegment"] for i in items if i.get("therapeuticSegment")})

    body = ",\n".join("  " + json.dumps(i, ensure_ascii=False) for i in items)
    seg_body = ",\n".join(
        f"  {{ id: {json.dumps(slug(s))}, label: {json.dumps(s)} }}" for s in segments
    )
    sec_body = ",\n".join(
        f"  {{ id: {json.dumps(i)}, label: {json.dumps(l)} }}" for i, l in SECTION_LABELS
    )

    ts = f"""// ─────────────────────────────────────────────
//  GENERATED FILE - do not edit by hand.
//  Regenerate:  python3 scripts/import_pharma.py
//  Source:      Jaydev Group Product list.xlsx
//
//  Pharmacopoeia grade, DMF, CEP and GMP status are NOT in the source
//  sheet and are deliberately absent here rather than guessed. The UI
//  renders "Grade & documentation on request" until they are supplied.
// ─────────────────────────────────────────────

export type PharmaSection = 'apis' | 'intermediates' | 'ingredients' | 'excipients' | 'others';
export type PharmaProduct = {{
  id: string;
  /** Full chemical name. Always searchable, even when shortName is shown. */
  name: string;
  /** Trade name shown on cards, from data/pharma-short-names.json. */
  shortName?: string;
  section: PharmaSection;
  cas?: string;
  /** Set when the product is sold in more than one isomeric form. */
  casForms?: {{ label: string; cas: string }}[];
  /** APIs only - the therapeutic filter axis. */
  therapeuticSegment?: string;
  investigational?: boolean;
}};

/** What a card and the search index display. */
export const displayName = (p: PharmaProduct) => p.shortName ?? p.name;

export const PHARMA_SECTIONS: {{ id: PharmaSection; label: string }}[] = [
{sec_body},
];

export const THERAPEUTIC_SEGMENTS: {{ id: string; label: string }}[] = [
{seg_body},
];

export const pharmaProducts: PharmaProduct[] = [
{body},
];

export const pharmaById = (id: string) => pharmaProducts.find((p) => p.id === id);
"""
    OUT.write_text(ts)

    counts = {}
    for i in items:
        counts[i["section"]] = counts.get(i["section"], 0) + 1
    print(f"wrote {OUT.relative_to(OUT.parent.parent)}  ({len(items)} items)")
    for sec, label in SECTION_LABELS:
        print(f"  {label:32} {counts.get(sec, 0):4}")
    print(f"  {'with CAS':32} {sum(1 for i in items if i.get('cas')):4}")
    print(f"  {'segmented (APIs)':32} {sum(1 for i in items if i.get('therapeuticSegment')):4}")
    print(f"  {'with short name':32} {sum(1 for i in items if i.get('shortName')):4}")
    print(f"  {'investigational':32} {sum(1 for i in items if i.get('investigational')):4}")
    print(f"  {'therapeutic segments':32} {len(segments):4}")
    missing = [i["name"] for i in items if not i.get("cas")]
    print(f"  {'still missing CAS':32} {len(missing):4}")
    for n in missing:
        print(f"      - {n}")
    if bad_cas:
        print("\n  !! FAILED CAS CHECK DIGIT - not emitted:")
        for n, c in bad_cas:
            print(f"      - {n}: {c}")
        sys.exit(1)


if __name__ == "__main__":
    main()
